from openpyxl import Workbook
from openpyxl import load_workbook  # 对xlsx文件的读写


def outxlsx(listt, title, name):  # 将列表listt输出到"{name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(title)
    for i in listt:
        ws.append(i)
    wb.save(name+".xlsx")


class animeIDs:

    def __init__(self, file) -> None:
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb.active
